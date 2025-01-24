
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime

from entity_app.domain.models.solicity import Status
from entity_app.adapters.serializers import NumeralResponseSerializer, SolicityResponseSerializer
from entity_app.domain.services.solicity_service import SolicityService
from entity_app.domain.services.transparency_active_service import TransparencyActiveService
from entity_app.domain.services.numeral_service import NumeralService
from entity_app.domain.services.transparency_colaborative_service import TransparencyColaborativeService
from entity_app.domain.services.transparency_focus_service import TransparencyFocusService

class ReportService:
    
    def __init__(self, solicity_service, transparency_service, numera_service,\
        transparency_collab,transparency_focus):
        self.solicity_service:SolicityService = solicity_service
        self.transparency_service:TransparencyActiveService = transparency_service
        self.numera_service:NumeralService = numera_service
        self.transparency_collab:TransparencyColaborativeService = transparency_collab
        self.transparency_focus: TransparencyFocusService = transparency_focus
        
        
        
        
    def generate_solicity_receiver(self,user_id:int,year:int):
        output_serializer_class = SolicityResponseSerializer

        solicity = self.solicity_service.get_entity_user_solicities(user_id)

        solicity = solicity.filter(status=Status.SEND)
        solicity = solicity.filter(created_at__year=year)
        # Create an Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Solicitudes Recibidas"

        # Write header row
        headers = ['N°','No. SAIP','Fecha Envío', 'Días Transcurridos', 'Estado']
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws[f'{column_letter}1'] = column_title

        lista_solicity = []
        # Write data rows
        for row_num, row_data in enumerate(solicity):
            row_ = {
                'index': row_num,
                'no_saip': row_data.number_saip,
                'created_at': row_data.created_at.strftime('%Y-%m-%d'),
                'days': (datetime.now().date() - row_data.created_at.date()).days,
                'status': row_data.status
            }
            
            lista_solicity.append(row_)
            
        for row_num, row_data in enumerate(lista_solicity, 2):
            for col_num, (column_name, cell_value) in enumerate(row_data.items(), 1):
                column_letter = get_column_letter(col_num)
                ws[f'{column_letter}{row_num}'] = cell_value
            

        # Save the workbook to a BytesIO stream
        from io import BytesIO
        response_stream = BytesIO()
        wb.save(response_stream)
        response_stream.seek(0)
        
        return response_stream
    
    def generate_all_solicities(self, user_id: int, year: int):
        output_serializer_class = SolicityResponseSerializer

        # Obtener las solicitudes del usuario para el año específico
        solicity = self.solicity_service.get_entity_user_solicities(user_id)
        #solicity = solicity.filter(created_at__year=year)  

        # Crear un libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Todas las Solicitudes"

        # Encabezados
        headers = ['N°', 'No. SAIP', 'Solicitante', 'Fecha Envío', 'Fecha Respuesta', 'Estado']
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws[f'{column_letter}1'] = column_title

        # Recopilar datos
        lista_solicity = []
        for row_num, row_data in enumerate(solicity):
            row_ = {
                'index': row_num + 1,  # Índice inicia en 1
                'no_saip': row_data.number_saip,
                'applicant': row_data.first_name + ' ' + row_data.last_name,
                'created_at': row_data.created_at.strftime('%Y-%m-%d'),
                'updated_at': row_data.updated_at.strftime('%Y-%m-%d'),
                'status': row_data.get_status_display(), # Estado en espanish
            }
            lista_solicity.append(row_)

        # Escribir datos en el Excel
        for row_num, row_data in enumerate(lista_solicity, 2):  # Inicia en la fila 2
            for col_num, (column_name, cell_value) in enumerate(row_data.items(), 1):
                column_letter = get_column_letter(col_num)
                ws[f'{column_letter}{row_num}'] = cell_value

        # Guardar el archivo en un stream
        from io import BytesIO
        response_stream = BytesIO()
        wb.save(response_stream)
        response_stream.seek(0)

        return response_stream
    
    
    def generate_solicities_response(self,user_id:int,year:int):
        output_serializer_class = SolicityResponseSerializer

        solicity = self.solicity_service.get_entity_user_solicities(user_id)

        solicity = solicity.filter(status__in=[Status.RESPONSED])
      
        #solicity = solicity.filter(created_at__year=year)
        # Create an Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Solicitudes Respondidas"

        # Write header row
        headers = ['N°', 'No. SAIP', 'Solicitante', 'Fecha Envío', 'Fecha Respuesta']
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws[f'{column_letter}1'] = column_title

        lista_solicity = []
        # Write data rows
        for row_num, row_data in enumerate(solicity):
            row_ = {
                'index': row_num,
                'no_saip': row_data.number_saip,
                'applicant': row_data.first_name + ' ' + row_data.last_name,
                'created_at': row_data.created_at.strftime('%Y-%m-%d'),
                'updated_at': row_data.updated_at.strftime('%Y-%m-%d'),
            }

            lista_solicity.append(row_)

        for row_num, row_data in enumerate(lista_solicity, 2):
            for col_num, (column_name, cell_value) in enumerate(row_data.items(), 1):
                column_letter = get_column_letter(col_num)
                ws[f'{column_letter}{row_num}'] = cell_value

        # Save the workbook to a BytesIO stream
        from io import BytesIO
        response_stream = BytesIO()
        wb.save(response_stream)
        response_stream.seek(0)
        
        return response_stream
    
    def generate_solicities_not_response(self, user_id: int,year:int):
        output_serializer_class = SolicityResponseSerializer

        solicity = self.solicity_service.get_entity_user_solicities(user_id)

        solicity = solicity.filter(
            status__in=[Status.NO_RESPONSED, Status.INFORMAL_MANAGMENT_NO_RESPONSED, Status.INSISTENCY_NO_RESPONSED])
        
        #solicity = solicity.filter(created_at__year=year)
        # Create an Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Solicitudes No Respondidas"

        # Write header row
        headers = ['N°', 'No. SAIP', 'Solicitante',
                   'Fecha Envío', 'Días Transcurridos']
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws[f'{column_letter}1'] = column_title

        lista_solicity = []
        # Write data rows
        for row_num, row_data in enumerate(solicity):
            # Determinar la fecha de fin (end_date) según el estado de la solicitud
            if row_data.status in ["RESPONSED", "INSISTENCY_RESPONSED"]:
                end_date = row_data.updated_at
            elif row_data.status in ["PRORROGA", "NO_RESPONSED", "INSISTENCY_NO_RESPONSED"]:
                end_date = row_data.expiry_date
            else:
                end_date = datetime.now()  # Fecha actual para estados no especificados

            # Asegurarse de que las fechas sean válidas
            if row_data.created_at and end_date:
                # Calcular la diferencia en días y horas
                diff_in_seconds = (end_date - row_data.created_at).total_seconds()
                days = diff_in_seconds // (24 * 3600)
                # En caso de que se requiera ser mas exacto
                # hours = (diff_in_seconds % (24 * 3600)) // 3600
            else:
                days = 0
                hours = 0

            # Construir el resultado para cada solicitud
            row_ = {
                'index': row_num,
                'no_saip': row_data.number_saip,
                'applicant': row_data.first_name + ' ' + row_data.last_name,
                'created_at': row_data.created_at.strftime('%Y-%m-%d') if row_data.created_at else 'N/A',
                'days': int(days),  # Convertir a entero
            }

            lista_solicity.append(row_)

        for row_num, row_data in enumerate(lista_solicity, 2):
            for col_num, (column_name, cell_value) in enumerate(row_data.items(), 1):
                column_letter = get_column_letter(col_num)
                ws[f'{column_letter}{row_num}'] = cell_value

        # Save the workbook to a BytesIO stream
        from io import BytesIO
        response_stream = BytesIO()
        wb.save(response_stream)
        response_stream.seek(0)

        return response_stream

    
    
    def generate_trasparency_report(self, est_id:int,year:int):

        
        
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Archivo de Transparencia"

        # Write header row
        headers = ['No.','Mes','Tipo de Transparencia','Descripción de Transparencia','Enlace a Archivo']
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws[f'{column_letter}1'] = column_title

        
        ta = self.transparency_service.get_by_year(year, est_id)
        tc = self.transparency_collab.get_by_year(year, est_id)
        tf = self.transparency_focus.get_by_year(year, est_id)
        
        list_final = []
        
        for row_num, row_data in enumerate(ta):
            for i in row_data.files.all():
                _row_data = {
                    'index': row_num,
                    'mes': row_data.month,
                    'tipo': 'Activa',
                    'descripcion': row_data.numeral.description,
                    'enlace':'https:transparencia.ec/v1/transparency'+ i.relative_url
                }

                list_final.append(_row_data)

        for row_num, row_data in enumerate(tc):
            for i in row_data.files.all():
                _row_data = {
                    'index': row_num,
                    'mes': row_data.month,
                    'tipo': 'Colaborativa',
                    'descripcion': 'Colaborativa',
                    'enlace': 'https:transparencia.ec/v1/transparency' + i.relative_url
                }

                list_final.append(_row_data)
                
                
        for row_num, row_data in enumerate(tf):
            for i in row_data.files.all():
                _row_data = {
                    'index': row_num,
                    'mes': row_data.month,
                    'tipo': 'Focalizada',
                    'descripcion': 'Focalizada',
                    'enlace': 'https:transparencia.ec/v1/transparency/' + i.relative_url
                }

                list_final.append(_row_data)
                
        for row_num, row_data in enumerate(list_final, 2):
            for col_num, (column_name, cell_value) in enumerate(row_data.items(), 1):
                column_letter = get_column_letter(col_num)
                ws[f'{column_letter}{row_num}'] = cell_value
            
        

        # Save the workbook to a BytesIO stream
        from io import BytesIO
        response_stream = BytesIO()
        wb.save(response_stream)
        response_stream.seek(0)

        return response_stream
