from openpyxl import Workbook
from django.db.models import Prefetch
from datetime import datetime
from entity_app.ports.repositories.anual_report_reposity import AnualReportReposity
import openpyxl
from openpyxl.styles import Alignment, Font
from entity_app.domain.models.establishment import EstablishmentExtended
from entity_app.domain.models.transparecy_colab import TransparencyColab
from entity_app.domain.models.transparecy_foc import TransparencyFocal
from entity_app.domain.models.transparency_active import EstablishmentNumeral, TransparencyActive
from entity_app.domain.models.solicity import Solicity, Status
from django.conf import settings
import os
from entity_app.domain.models.anual_report import GeneralAnualReport, GenerateAnualReport
from entity_app.domain.models.pnt1 import Pnt1_Active, Pnt1_Colab, Pnt1_Focal, Pnt1_Pasive
import re


class AnualReportService:
    def __init__(self, anual_report_repository: AnualReportReposity):
        self.anual_report_repository = anual_report_repository

    def create(self, anual_report: dict):
        return self.anual_report_repository.add(**anual_report)

    def get(self, establishment_id: int, year: int, ):
        return self.anual_report_repository.get(establishment_id, year)

    def get_all(self):
        return self.anual_report_repository.get_all()

    def update(self, anual_report: dict):
        return self.anual_report_repository.update(**anual_report)

    def delete(self, anual_report_id):
        return self.anual_report_repository.delete(anual_report_id)

    def generate_unique(self, year, establishment_id, update_state):

        try:

            reporte_existente = GenerateAnualReport.objects.filter(
                establishment_id=establishment_id, year=year).first()
            if reporte_existente:
                reporte_existente.delete()

            meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
            establishments = EstablishmentExtended.objects.filter(is_active=True, id=establishment_id).prefetch_related(
                Prefetch('transparency_colab', queryset=TransparencyColab.objects.filter(
                    year=year), to_attr='tc'),
                Prefetch('transparency_focal', queryset=TransparencyFocal.objects.filter(
                    year=year), to_attr='tf'),
                Prefetch('transparency_active', queryset=TransparencyActive.objects.filter(
                    year=year), to_attr='ta')
            ).order_by('name')

            solicities = Solicity.objects.filter(
                deleted=False, date__year=year,
                establishment_id=establishment_id)

            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "T.A-F-C"

            # Configuración de columnas y anchos
            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G',
                       'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
            widths = [30, 30, 30, 30, 10, 10, 10,
                      10, 10, 10, 10, 10, 10, 10, 10, 10]
            for col, width in zip(columns, widths):
                ws.column_dimensions[col].width = width

            # Encabezados de la hoja
            ws.append([
                "Función a la que pertenece", "Tipo", "Nombre Entidad", "Marco Legal",
                "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto",
                "Septiembre", "Octubre", "Noviembre", "Diciembre"
            ])

            # Estilo de las celdas
            for cell in ws["1:1"]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    wrap_text=True, vertical='center', horizontal='center')

            # Funciones auxiliares

            def append_row(ws, data, bold_col=4):
                ws.append(data)
                ws.cell(row=ws.max_row, column=bold_col).font = Font(bold=True)
                ws.cell(row=ws.max_row, column=bold_col).alignment = Alignment(
                    wrap_text=True, vertical='center', horizontal='center')

            def append_row_without_bold(ws, data):
                ws.append(data)
                ws.cell(row=ws.max_row, column=4).alignment = Alignment(
                    wrap_text=True, vertical='center', horizontal='center')

            # Procesar datos de transparencia para cada establecimiento
            for number, establishment in enumerate(establishments):
                ta_pnt1 = Pnt1_Active.objects.filter(
                    identification=establishment.identification)

                tc_pnt1 = Pnt1_Colab.objects.filter(
                    identification=establishment.identification)

                tf_pnt1 = Pnt1_Focal.objects.filter(
                    identification=establishment.identification)

                # Transparencia colaborativa
                function_type = establishment.function_organization.name if establishment.function_organization else ""
                progress = (number+1) / len(establishments) * 100
                progress = round(progress, 2)
                update_state(state='PROGRESS', meta={
                    'progress': 50,
                    'message': f"Procesando publicaciones de transparencia... Obteniendo información  | Creando parte 1/2"

                })
                if hasattr(establishment, 'tc') and len(establishment.tc) > 0:

                    list_transparency = {month: False for month in meses}
                    function_type = establishment.function_organization.name if establishment.function_organization else ""

                    for t in establishment.tc:
                        if 1 <= t.month <= 12:
                            list_transparency[meses[t.month-1]] = True
                    if year == 2024:
                        for t in tc_pnt1:
                            if t.enero:
                                list_transparency[meses[0]] = True
                            if t.febrero:
                                list_transparency[meses[1]] = True
                            if t.marzo:
                                list_transparency[meses[2]] = True
                            if t.abril:
                                list_transparency[meses[3]] = True
                            if t.mayo:
                                list_transparency[meses[4]] = True
                            if t.junio:
                                list_transparency[meses[5]] = True
                            if t.julio:
                                list_transparency[meses[6]] = True
                            if t.agosto:
                                list_transparency[meses[7]] = True
                    append_row(ws, [
                        function_type, "Pública", establishment.name, "Art. 4, número 9 T. Colaborativa",
                        *['si' if list_transparency[month] else 'no' for month in meses]
                    ])
                else:
                    append_row(ws, [function_type, "Pública", establishment.name,
                                    "Art. 4, número 9 T. Colaborativa"] + ['no'] * 12)

                # Transparencia focalizada
                if hasattr(establishment, 'tf') and len(establishment.tf) > 0:
                    list_transparency = {month: False for month in meses}
                    function_type = establishment.function_organization.name if establishment.function_organization else ""

                    for t in establishment.tf:
                        if 1 <= t.month <= 12:
                            list_transparency[meses[t.month-1]] = True

                    if year == 2024:
                        for t in tf_pnt1:
                            if t.enero:
                                list_transparency[meses[0]] = True
                            if t.febrero:
                                list_transparency[meses[1]] = True
                            if t.marzo:
                                list_transparency[meses[2]] = True
                            if t.abril:
                                list_transparency[meses[3]] = True
                            if t.mayo:
                                list_transparency[meses[4]] = True
                            if t.junio:
                                list_transparency[meses[5]] = True
                            if t.julio:
                                list_transparency[meses[6]] = True
                            if t.agosto:
                                list_transparency[meses[7]] = True

                    append_row(ws, [
                        function_type, "Pública", establishment.name, "Art. 4 número 10 T. Focalizada", *
                        ['si' if list_transparency[month]
                            else 'no' for month in meses]
                    ])
                else:
                    append_row(ws, [function_type, "Pública", establishment.name,
                                    "Art. 4 número 10 T. Focalizada"] + ['no'] * 12)

                # Transparencia activa
                list_check_ta = []

                if hasattr(establishment, 'ta') and len(establishment.ta) > 0:

                    publications = [
                        ta for ta in establishment.ta if ta.numeral.is_default]
                    publications = sorted(
                        publications, key=lambda x: x.numeral.name)
                    for publication in publications:
                        list_check_ta.append({
                            'numeral': publication.numeral.name,
                            'month': publication.month,
                            'published': publication.published
                        })

                numerales_asignados = EstablishmentNumeral.objects.filter(
                    establishment=establishment, numeral__is_default=True).order_by('numeral__name')
                append_row(ws, [function_type, "Pública",
                                establishment.name, 'Art. 19 T. Activa'] + [''] * 12)
                
                # Numerales asignados para Art. 19 T. Activa
                for i in numerales_asignados:
                    mes_checks = ['si' if any(x['numeral'] == i.numeral.name and x['month'] == _x +
                                              1 and x['published'] for x in list_check_ta) else 'no' for _x in range(12)]

                    if year == 2024:
                        numero = re.search(
                            r'\d+(\.\d+)?(-\d+)?', i.numeral.name)
                        numero = numero.group() if numero else ""
                        numero = numero.replace("-", " - ")
                        ta = ta_pnt1.filter(numeral=numero,
                                            art="19"
                                            )

                        for _ta in ta:
                            if _ta.enero:
                                mes_checks[0] = 'si'
                            if _ta.febrero:
                                mes_checks[1] = 'si'
                            if _ta.marzo:
                                mes_checks[2] = 'si'
                            if _ta.abril:
                                mes_checks[3] = 'si'
                            if _ta.mayo:
                                mes_checks[4] = 'si'
                            if _ta.junio:
                                mes_checks[5] = 'si'
                            if _ta.julio:
                                mes_checks[6] = 'si'
                            if _ta.agosto:
                                mes_checks[7] = 'si'

                    append_row_without_bold(ws, [function_type, "Pública",
                                            establishment.name, i.numeral.name.replace('Numeral', ''), *mes_checks])

                numerales_asignados_esp = EstablishmentNumeral.objects.filter(
                    establishment=establishment, numeral__is_default=False).order_by('numeral__name')
                publications = [
                    ta for ta in establishment.ta if not ta.numeral.is_default]
                publications = sorted(
                    publications, key=lambda x: x.numeral.name)
                for publication in publications:
                    list_check_ta.append({
                        'numeral': publication.numeral.name,
                        'month': publication.month,
                        'published': publication.published
                    })
                if len(numerales_asignados_esp) > 0:
                    '''append_row(ws, [function_type, "Pública",
                                    establishment.name, 'Obligaciones Específicas'] + [''] * 12)'''
                    for i in numerales_asignados_esp:
                        mes_checks = ['si' if any(x['numeral'] == i.numeral.name and x['month'] == _x +
                                                  1 and x['published'] for x in list_check_ta) else 'no' for _x in range(12)]

                        if year == 2024:
                            numero = re.search(
                                r'\d+(\.\d+)?(-\d+)?', i.numeral.name)
                            numero = numero.group() if numero else ""
                            numero = numero.replace("-", " - ")
                            ta = ta_pnt1.filter(
                                numeral=numero
                            ).exclude(art="19")

                            for _ta in ta:
                                if _ta.enero:
                                    mes_checks[0] = 'si'
                                if _ta.febrero:
                                    mes_checks[1] = 'si'
                                if _ta.marzo:
                                    mes_checks[2] = 'si'
                                if _ta.abril:
                                    mes_checks[3] = 'si'
                                if _ta.mayo:
                                    mes_checks[4] = 'si'
                                if _ta.junio:
                                    mes_checks[5] = 'si'
                                if _ta.julio:
                                    mes_checks[6] = 'si'
                                if _ta.agosto:
                                    mes_checks[7] = 'si'

                        append_row_without_bold(ws, [function_type, "Pública",
                                                establishment.name, "Obligaciones Específicas: "+i.numeral.name.replace('Numeral', ''), *mes_checks])

            # Crear nueva hoja "Pasiva"
            ws = wb.create_sheet(title="Pasiva")
            ws.append([
                "Función a la que pertenece", "Tipo", "Nombre Entidad",
                "No. Solicitud", "Solicitante", "Fecha de envio", "Fecha Respuesta",
                "Estado"
            ])

            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            widths = [30, 30, 30, 30, 30, 30, 30, 30]
            for col, width in zip(columns, widths):
                ws.column_dimensions[col].width = width

            # Procesar solicitudes
            for number, establishment in enumerate(establishments):
                progress_calculate = (number+1)
                progress = progress_calculate / len(establishments) * 100
                progress = round(progress, 2)

                all_solicities = solicities.filter(
                    establishment=establishment)
                function_type = establishment.function_organization.name if establishment.function_organization else ""

                for x, solicity in enumerate(all_solicities):

                    progress = progress_calculate / len(all_solicities) * 100
                    progress = progress / 2
                    update_state(state='PROGRESS', meta={
                        'progress': 50+progress,
                        'message': f"Procesando solicitudes... Obteniendo información | Creando parte 2/2"
                    })
                    response = solicity.solicityresponse_set.first()
                    if response and response.created_at:
                        date = response.created_at.strftime("%d/%m/%Y")

                    else:
                        date = ""
                    solicity_date = solicity.date.strftime("%d/%m/%Y")
                    status = Status(solicity.status).label
                    append_row_without_bold(ws, [
                        function_type, "Pública", solicity.establishment.name,
                        solicity.number_saip, f"{solicity.first_name} {
                            solicity.last_name}", solicity_date, date, status
                    ])
                if year == 2024:
                    solicities_pnt1 = Pnt1_Pasive.objects.filter(
                        identification=establishment.identification)
                    for solicity in solicities_pnt1:
                        append_row_without_bold(ws, [
                            function_type, "Pública", establishment.name,
                            solicity.saip, solicity.name_solicitant, solicity.date, solicity.date_response, solicity.state
                        ])

            # Guardar archivo
            report_name = f'reporte_anual_{year}_{
                datetime.now().strftime("%d-%m-%Y")}.xlsx'
            path = os.path.join(settings.MEDIA_ROOT, 'transparencia', establishment.identification, str(year),
                                'reporte_anual')
            if not os.path.exists(path):
                os.makedirs(path)

            path = os.path.join(path, report_name)

            wb.save(path)
            GenerateAnualReport.objects.create(
                establishment_id=establishment_id, year=year,
                file=path.replace('code/media/', '')
            )
            return {'path': path}
        except Exception as e:
            print(e)
            return {'error': str(e)}

    def generate(self, year, update_state):
        try:
            if type(year) == str:
                year = int(year)

            meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
            establishments = EstablishmentExtended.objects.filter(is_active=True).prefetch_related(
                Prefetch('transparency_colab', queryset=TransparencyColab.objects.filter(
                    year=year), to_attr='tc'),
                Prefetch('transparency_focal', queryset=TransparencyFocal.objects.filter(
                    year=year), to_attr='tf'),
                Prefetch('transparency_active', queryset=TransparencyActive.objects.filter(
                    year=year), to_attr='ta')
            ).order_by('name')

            solicities = Solicity.objects.filter(
                deleted=False, date__year=year)
            pasive_pnt1 = Pnt1_Pasive.objects.all()
            active_pnt1 = Pnt1_Active.objects.all()
            colab_pnt1 = Pnt1_Colab.objects.all()
            focal_pnt1 = Pnt1_Focal.objects.all()
            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "T.A-F-C"

            # Configuración de columnas y anchos
            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G',
                       'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
            widths = [30, 30, 30, 30, 10, 10, 10,
                      10, 10, 10, 10, 10, 10, 10, 10, 10]
            for col, width in zip(columns, widths):
                ws.column_dimensions[col].width = width

            # Encabezados de la hoja
            ws.append([
                "Función a la que pertenece", "Tipo", "Nombre Entidad", "Marco Legal",
                "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto",
                "Septiembre", "Octubre", "Noviembre", "Diciembre"
            ])

            # Estilo de las celdas
            for cell in ws["1:1"]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    wrap_text=True, vertical='center', horizontal='center')

            # Funciones auxiliares

            def append_row(ws, data, bold_col=4):
                ws.append(data)
                ws.cell(row=ws.max_row, column=bold_col).font = Font(bold=True)
                ws.cell(row=ws.max_row, column=bold_col).alignment = Alignment(
                    wrap_text=True, vertical='center', horizontal='center')

            def append_row_without_bold(ws, data):
                ws.append(data)
                ws.cell(row=ws.max_row, column=4).alignment = Alignment(
                    wrap_text=True, vertical='center', horizontal='center')

            def process_transparency(ws, establishment, transparency, legal_text, months):
                list_transparency = {month: False for month in months}
                function_type = establishment.function_organization.name if establishment.function_organization else ""

                for t in transparency:
                    if 1 <= t.month <= 12:
                        list_transparency[months[t.month-1]] = True
                append_row(ws, [
                    function_type, "Pública", establishment.name, legal_text,
                    *['si' if list_transparency[month] else 'no' for month in months]
                ])

            # Procesar datos de transparencia para cada establecimiento
            for number, establishment in enumerate(establishments):
                # Transparencia colaborativa
                function_type = establishment.function_organization.name if establishment.function_organization else ""
                progress = (number+1) / len(establishments) * 100
                progress = round(progress, 2)
                update_state(state='PROGRESS', meta={
                    'progress': progress,
                    'message': f"Procesando publicaciones de transparencia... Obteniendo información {number+1}/{len(establishments)} entidades... | Creando parte 1/2"

                })
                if hasattr(establishment, 'tc') and len(establishment.tc) > 0:

                    list_transparency = {month: False for month in meses}
                    function_type = establishment.function_organization.name if establishment.function_organization else ""

                    for t in establishment.tc:
                        if 1 <= t.month <= 12:
                            list_transparency[meses[t.month-1]] = True

                    if year == 2024:
                        tc = colab_pnt1.filter(
                            identification=establishment.identification
                        )
                        if tc:
                            for t in tc:
                                if t.enero:
                                    list_transparency[meses[0]] = True
                                if t.febrero:
                                    list_transparency[meses[1]] = True
                                if t.marzo:
                                    list_transparency[meses[2]] = True
                                if t.abril:
                                    list_transparency[meses[3]] = True
                                if t.mayo:
                                    list_transparency[meses[4]] = True
                                if t.junio:
                                    list_transparency[meses[5]] = True
                                if t.julio:
                                    list_transparency[meses[6]] = True
                                if t.agosto:
                                    list_transparency[meses[7]] = True

                    append_row(ws, [
                        function_type, "Pública", establishment.name, "Art. 4, número 9 T. Colaborativa",
                        *['si' if list_transparency[month] else 'no' for month in meses]
                    ])
                else:
                    append_row(ws, [function_type, "Pública", establishment.name,
                                    "Art. 4, número 9 T. Colaborativa"] + ['no'] * 12)

                # Transparencia focalizada
                if hasattr(establishment, 'tf') and len(establishment.tf) > 0:
                    process_transparency(ws, establishment, establishment.tf,
                                         "Art. 4 número 10 T. Focalizada", meses)
                    list_transparency = {month: False for month in meses}
                    function_type = establishment.function_organization.name if establishment.function_organization else ""

                    for t in establishment.tf:
                        if 1 <= t.month <= 12:
                            if year == 2024:
                                list_transparency[meses[t.month-1]] = True
                    if year == 2024:
                        tf = focal_pnt1.filter(
                            identification=establishment.identification
                        )
                        for t in tf:
                            if t.enero:
                                list_transparency[meses[0]] = True
                            if t.febrero:
                                list_transparency[meses[1]] = True
                            if t.marzo:
                                list_transparency[meses[2]] = True
                            if t.abril:
                                list_transparency[meses[3]] = True
                            if t.mayo:
                                list_transparency[meses[4]] = True
                            if t.junio:
                                list_transparency[meses[5]] = True
                            if t.julio:
                                list_transparency[meses[6]] = True
                            if t.agosto:
                                list_transparency[meses[7]] = True
                    append_row(ws, [
                        function_type, "Pública", establishment.name, "Art. 4, número 10 T. Focalizada",
                        *['si' if list_transparency[month] else 'no' for month in meses]
                    ])

                else:
                    append_row(ws, [function_type, "Pública", establishment.name,
                                    "Art. 4 número 10 T. Focalizada"] + ['no'] * 12)

                # Transparencia activa
                list_check_ta = []

                if hasattr(establishment, 'ta') and len(establishment.ta) > 0:

                    publications = [
                        ta for ta in establishment.ta if ta.numeral.is_default]
                    publications = sorted(
                        publications, key=lambda x: x.numeral.name)
                    for publication in publications:
                        list_check_ta.append({
                            'numeral': publication.numeral.name,
                            'month': publication.month,
                            'published': publication.published
                        })

                numerales_asignados = EstablishmentNumeral.objects.filter(
                    establishment=establishment, numeral__is_default=True).order_by('numeral__name')
                append_row(ws, [function_type, "Pública",
                                establishment.name, 'Art. 19 T. Activa'] + [''] * 12)

                # Numerales asignados para Art. 19 T. Activa
                for i in numerales_asignados:

                    mes_checks = ['si' if any(x['numeral'] == i.numeral.name and x['month'] == _x +
                                              1 and x['published'] for x in list_check_ta) else 'no' for _x in range(12)]

                    if year == 2024:
                        numero = re.search(
                            r'\d+(\.\d+)?(-\d+)?', i.numeral.name)
                        numero = numero.group() if numero else ""
                        numero = numero.replace("-", " - ")
                        ta = active_pnt1.filter(
                            identification=establishment.identification,
                            numeral=numero,
                            art="19"
                        )

                        for _ta in ta:
                            if _ta.enero:
                                mes_checks[0] = 'si'
                            if _ta.febrero:
                                mes_checks[1] = 'si'
                            if _ta.marzo:
                                mes_checks[2] = 'si'
                            if _ta.abril:
                                mes_checks[3] = 'si'
                            if _ta.mayo:
                                mes_checks[4] = 'si'
                            if _ta.junio:
                                mes_checks[5] = 'si'
                            if _ta.julio:
                                mes_checks[6] = 'si'
                            if _ta.agosto:
                                mes_checks[7] = 'si'

                    append_row_without_bold(ws, [function_type, "Pública",
                                            establishment.name, i.numeral.name.replace('Numeral', ''), *mes_checks])

                numerales_asignados_esp = EstablishmentNumeral.objects.filter(
                    establishment=establishment, numeral__is_default=False).order_by('numeral__name')
                
                
                publications = [
                    ta for ta in establishment.ta if not ta.numeral.is_default]
                publications = sorted(
                    publications, key=lambda x: x.numeral.name)
                for publication in publications:
                    list_check_ta.append({
                        'numeral': publication.numeral.name,
                        'month': publication.month,
                        'published': publication.published
                    })
                    
                    
                    
                if len(numerales_asignados_esp) > 0:

                    '''append_row(ws, [function_type, "Pública",
                                    establishment.name, 'Obligaciones Específicas'] + [''] * 12)'''
                    for i in numerales_asignados_esp:
                        mes_checks = ['si' if any(x['numeral'] == i.numeral.name and x['month'] == _x +
                                                  1 and x['published'] for x in list_check_ta) else 'no' for _x in range(12)]
                        if year == 2024:
                            numero = re.search(
                                r'\d+(\.\d+)?(-\d+)?', i.numeral.name)
                            numero = numero.group() if numero else ""
                            numero = numero.replace("-", " - ")
                            ta = active_pnt1.filter(
                                identification=establishment.identification,
                                numeral=numero
                            ).exclude(art="19")

                            for _ta in ta:
                                if _ta.enero:
                                    mes_checks[0] = 'si'
                                if _ta.febrero:
                                    mes_checks[1] = 'si'
                                if _ta.marzo:
                                    mes_checks[2] = 'si'
                                if _ta.abril:
                                    mes_checks[3] = 'si'
                                if _ta.mayo:
                                    mes_checks[4] = 'si'
                                if _ta.junio:
                                    mes_checks[5] = 'si'
                                if _ta.julio:
                                    mes_checks[6] = 'si'
                                if _ta.agosto:
                                    mes_checks[7] = 'si'

                        append_row_without_bold(ws, [function_type, "Pública",
                                                establishment.name, "Obligaciones Específicas: " + i.numeral.name.replace('Numeral', ''), *mes_checks])

            # Crear nueva hoja "Pasiva"
            ws = wb.create_sheet(title="Pasiva")
            ws.append([
                "Función a la que pertenece", "Tipo", "Nombre Entidad",
                "No. Solicitud", "Solicitante", "Fecha de envio", "Fecha Respuesta",
                "Estado"
            ])

            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            widths = [30, 30, 30, 30, 30, 30, 30, 30]
            for col, width in zip(columns, widths):
                ws.column_dimensions[col].width = width

            # Procesar solicitudes
            for number, establishment in enumerate(establishments):
                progress_calculate = (number+1)
                progress = progress_calculate / len(establishments) * 100
                progress = round(progress, 2)

                update_state(state='PROGRESS', meta={
                    'progress': progress,
                    'message': f"Procesando solicitudes... Obteniendo información {progress_calculate}/{len(establishments)} entidades... | Creando parte 2/2"
                })
                all_solicities = solicities.filter(
                    establishment=establishment)
                function_type = establishment.function_organization.name if establishment.function_organization else ""

                for solicity in all_solicities:
                    response = solicity.solicityresponse_set.first()
                    if response and response.created_at:
                        date = response.created_at.strftime("%d/%m/%Y")

                    else:
                        date = ""
                    solicity_date = solicity.date.strftime("%d/%m/%Y")
                    status = Status(solicity.status).label
                    append_row_without_bold(ws, [
                        function_type, "Pública", solicity.establishment.name,
                        solicity.number_saip, f"{solicity.first_name} {
                            solicity.last_name}", solicity_date, date, status
                    ])

                if year == 2024:
                    solicities_pnt1 = pasive_pnt1.filter(
                        identification=establishment.identification)
                    for solicity in solicities_pnt1:
                        append_row_without_bold(ws, [
                            function_type, "Pública", establishment.name,
                            solicity.saip, solicity.name_solicitant, solicity.date, solicity.date_response, solicity.state
                        ])

            # Guardar archivo
            report_name = f'reporte_anual_{year}_{
                datetime.now().strftime("%d-%m-%Y")}.xlsx'
            path = os.path.join(
                settings.MEDIA_ROOT, 'transparencia', 'reporte_anual_general', str(year))

            if not os.path.exists(path):
                os.makedirs(path)

            path = os.path.join(path, report_name)

            GeneralAnualReport.objects.filter(year=year).delete()

            GeneralAnualReport.objects.create(
                year=year,
                file=path.replace('code/media/', '')
            )

            wb.save(path)
            return {'path': path.replace('code/', ''), 'url': path.replace('code/', '')}
        except Exception as e:
            print(e)
            return {'error': str(e)}



    def template_anual_report_form(self):

        

        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Anual"
        
        ws.append([''])
        #ponerle color celeste de fondo a la celda con texto rojo
        #azul enfasis 5. 50% claro
        ws['A2'].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type = "solid")
        ws['A2'].font = Font(color="FF0000")
        ws['A2'] = "LITERAL A"
        #combinar celda de A1 a L1
        ws.merge_cells('A2:L2')


        ws.append(["Normativa","Número de meses publicados"])
        #poner negrita
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)

        ws.append(["Art. 4, número 9 T. Colaborativa",""])
        ws.append(["Art. 4, número 10 T. Focalizada",""])
        ws.append(["Art. 19 T. Activa",""])
        ws.append(["Obligaciones Específicas",""])

        ws.append([""])


        ws.append(["Artículo 10 de la LOTAIP"])
        ws['A9'].font = Font(bold=True)
        ws.append(['Custoria de la Información'])
        ws['A10'].font = Font(bold=True)
        ws.append(['¿Se ha creado y se mantiene registros públicos de manera profesional para el manejo y archivo de la información y documentación?'])
        ws.append([''])

        ws.append(['Descripción','Valor'])
        ws['A13'].font = Font(bold=True)
        ws['B13'].font = Font(bold=True)


        ws.append(['Si/No',''])
        ws.append(['Norma achivistica utilizada',''])
        ws.append(['Comentario/Aclaración',''])

        ws.append([''])
        ws.append(['Artículo 17 y 18 de la LOTAIP'])
        ws['A18'].font = Font(bold=True)
        ws.append(['Información reservada'])
        ws['A19'].font = Font(bold=True)
        ws.append([''])
        ws.append(['Descripción','Valor'])
        ws['A21'].font = Font(bold=True)
        ws['B21'].font = Font(bold=True)
        ws.append(['Si/No',''])
        ws.append([''])
        ws.append([''])


        ws.append(['Artículo 40 de la LOTAIP'])
        ws['A24'].font = Font(bold=True)
        ws.append(['Gestión Oficiosa'])
        ws['A25'].font = Font(bold=True)
        ws.append(['¿Alguna persona que solicitó información indicó que la recibida no era de calidad, o existió ambigüedad en el manejo de la información registrada en el Portal Nacional de Transparencia o sobre la información que se difunde en la propia institución, y resolvió solicitar la corrección en la difusión de la información; o alguna persona solicitó la intervención del Defensor del Pueblo para que se corrija y se brinde mayor claridad y sistematización en la organización de la información ?'])
        #combinar celdas de A20 a L30
        ws.merge_cells('A28:L30')


        ws.append(['Descripción','Valor'])
        ws['A31'].font = Font(bold=True)
        ws['B31'].font = Font(bold=True)
        ws.append(['Si/No',''])
        ws.append(['Cantidad',''])
        ws.append(['Descripción específica de la corrección de la información',''])
        ws.append([''])

        ws.append(['Artículo 42 de la LOTAIP'])
        ws['A35'].font = Font(bold=True)
        ws.append(['Sanciones administrativas'])
        ws['A36'].font = Font(bold=True)
        ws.append(['¿Personas servidoras públicas de su entidad o personas del sector privado han recibido sanciones por omisión o negativa en el acceso a la información pública?'])
        ws.append([''])
        ws.append(['Descripción','Valor'])
        ws['A39'].font = Font(bold=True)
        ws['B39'].font = Font(bold=True)
        ws.append(['Si/No',''])

        ws.append([''])
        ws.append(['Descripción','Cantidad','Descripción específica de la sanción administrativa'])
        ws['A42'].font = Font(bold=True)
        ws['B42'].font = Font(bold=True)
        ws['C42'].font = Font(bold=True)
        ws.append(['Ley Orgánica del Servicio Público','',''])
        ws.append(['Ley Orgánica de la Contraloría General del Estado','',''])
        ws.append(['Ley Orgánica del Sistema Nacional de Contratación Pública','',''])
        ws.append(['Ley Orgánica de Participación Ciudadana','',''])
        ws.append([''])
        ws.append([''])


        ws.append(['Disposición transitoria septima LOTAIP'])
        ws['A48'].font = Font(bold=True)
        ws.append(['¿Su entidad implementó programas de difusión, capacitacion y fortalecimiento sobre la LOTAIP dirigido a las personas servidoras públicas de su institución?'])
        ws.append(['Descripción','Valor'])
        ws['A51'].font = Font(bold=True)
        ws['B51'].font = Font(bold=True)
        ws.append(['Si/No',''])
        ws.append(['Cantidad',''])
        ws.append(['Descripción específica del programa de difusión, capacitación y fortalecimiento sobre la LOTAIP',''])

        ws.append([''])
        ws.append([''])
        ws.append([''])
        ws.append(['Disposición transitoria octava LOTAIP'])
        ws['A57'].font = Font(bold=True)
        ws.append(['¿Sí su entidad es un establecimiento educativo público o privado, desarrolló actividades y programas de promoción del derecho de acceso a la información pública, sus garantías y referente a la transparencia sus garantías y referente a la transparencia colaborativa?'])
        ws.append([''])
        ws.append(['Descripción','Valor'])
        ws['A60'].font = Font(bold=True)
        ws['B60'].font = Font(bold=True)
        ws.append(['Si/No',''])
        ws.append(['Cantidad',''])
        ws.append(['Descripción específica de la actividad o programa desarrollado',''])
        ws.append([''])


        ws.append(['LITERAL "B"'])
        
        ws.merge_cells('A70:L70')
        #ponerle color celeste de fondo a la celda con texto rojo
        #azul enfasis 5. 50% claro
        ws['A70'].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type = "solid")
        ws['A70'].font = Font(color="FF0000")
        ws.append([''])
        ws.append('Ingrese el número de solicitudes de acceso a la información pública que su entidad recibió y gestionó en el período enero-diciembre')
        ws.append([''])
        ws.append(['Descripción','Valor'])
        ws['A74'].font = Font(bold=True)
        ws['B74'].font = Font(bold=True)
        ws.append(['Cantidad',''])
        ws.append([''])
        ws.append(['¿Su entidad recibió y gestionó una cantidad  diferente de solicitudes de las que registro en el Portal Nacional de Transparencia?'])
        ws.append([''])


        ws.append(['Descripción','Valor'])
        ws['A79'].font = Font(bold=True)
        ws['B79'].font = Font(bold=True)
        ws.append(['Si/No',''])
        ws.append(['Cantidad de SAIP registradas en el portal',''])
        ws.append(['Cantidad de SAIP NO registradas en el portal',''])
        ws.append(['Descripción de las razones por las que no fueron ingresadas al portal',''])
        ws.append([''])
        ws.append([''])
        ws.append(['¿Las solicitudes de acceso a la información pública que no fueron registradas en el Portal Nacional de Transparencia, fueron respondidas?'])
        ws.append([''])
        ws.append(['Descripción','Valor'])
        ws['A87'].font = Font(bold=True)
        ws['B87'].font = Font(bold=True)
        ws.append(['Cantidad',''])
        ws.append(['Comentario/aclaración',''])
        ws.append([''])
        ws.append([''])
        ws.append([''])

        ws.append(['LITERAL "C"'])
        ws['A92'].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type = "solid")
        ws['A92'].font = Font(color="FF0000")
        ws.merge_cells('A95:L95')
        ws.append([''])
        ws.append(['c) Informe semestral actualizado sobre el listado índice de información reservada'])
        ws.append([''])
        ws.append(['Número de temas clasificados como:'])
        ws.append(['Descripción','Valor'])
        ws.append(['Reservado',''])
        ws.append(['Confidencial (Ley Organica de Empresas Públicas)',''])
        ws.append([''])
        ws.append([''])


        ws.append(['LITERAL "D"'])
        ws['A105'].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type = "solid")
        ws['A105'].font = Font(color="FF0000")
        ws.merge_cells('A105:L105')

        ws.append([''])
        ws.append(['d) El índice de la información clasificada como reservada, detallando la fecha de la resolución de clasificación de la reserva y el período de vigencia de la misma'])
        ws.append([''])
        
        ws.append(['Tema','Base Legal','Fecha de clasificación de la información reservada - semestral','Periodo de vigencia de la clasificación de la reserva','Se ha efectuado ampliación','Descripción de la ampliación','Fecha de la ampliación','Periodo de vigencia de a ampliación'])


